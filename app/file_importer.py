# -*- coding: utf-8 -*-
"""
@author: Graham.Macleod
"""

import pandas as pd

''' Provides classes for reading files. '''
class CSVReaderBehaviours:

    def read_file(file, **kwargs):
        try:
            file.seek(0)
            return pd.read_csv(file, **kwargs)
        except:
            return pd.DataFrame({})

    def get_buffer_size(file_path):
        return 1000

class ExcelReaderBehaviours:

    def read_file(file, **kwargs):
        try:
            file.seek(0)
            return pd.read_excel(file, **kwargs)
        except:
            return pd.DataFrame({})

    def get_buffer_size(file_path):
        ''' Uses openpyxl to get the number of rows in the excel file. 
            Uses the number of rows to make a guess a suitable buffer size.'''
        from openpyxl import load_workbook
        workbook = load_workbook(file_path, read_only=True)
        sheet = workbook.active
        row_count = sheet.max_row
        if row_count >= 20000:
            return 10000
        if row_count >= 2000:
            return 1000
        return row_count

    def get_sheet_names(self):
        from openpyxl import load_workbook
        workbook = load_workbook(self.file, read_only=True)
        return workbook.sheetnames
    
class FileReaderBehaviours:

    def read_detection_preview(self, *pargs, **kwargs):
        return self.filetype_behaviour.read_file(*pargs, nrows=10, **kwargs)

    def read_display_preview(self, *pargs, **kwargs):
        return self.filetype_behaviour.read_file(*pargs, nrows=10, dtype=str, **kwargs)

    def read_preview_with_datetime_column(self, file, dt_col, **kwargs):
        df = self.read_detection_preview(file, **kwargs)
        df.rename(columns={"datetime": "datetime_old"}, inplace=True)
        df.insert(0, 'datetime', pd.to_datetime(df.iloc[:, dt_col]))
        return df.drop(columns=df.columns[dt_col + 1])

    def read_preview_with_date_and_time_column(self, file, d_col, t_col, **kwargs):
        df = self.read_detection_preview(file, **kwargs)
        df.rename(columns={"datetime": "datetime_old"}, inplace=True)
        df.insert(0, 'datetime', pd.to_datetime(
            pd.to_datetime(df.iloc[:, d_col]).dt.date.astype(str) + 
            ' ' + df.iloc[:, t_col].astype(str)))
        return df.drop(columns=df.columns[[d_col + 1, t_col + 1]])

    def read_full(self, file, *dt_col_indexes, **kwargs):
        df = self.buffered_read_until_empty_row(file, **kwargs)
        if len(dt_col_indexes) == 1:
            dt_col = dt_col_indexes[0]
            df.rename(columns={"datetime": "datetime_old"}, inplace=True)
            df.insert(0, 'datetime', pd.to_datetime(df.iloc[:, dt_col]))
            df = df.drop(columns=df.columns[[dt_col + 1]])
        else:
            d_col, t_col = dt_col_indexes
            datetime_col = pd.to_datetime(
                           pd.to_datetime(df.iloc[:, d_col]).dt.date.astype(str) +
                           ' ' + df.iloc[:, t_col].astype(str))
            df.rename(columns={"datetime": "datetime_old"}, inplace=True)
            df.insert(0, 'datetime', datetime_col)
            return df.drop(columns=df.columns[[d_col + 1, t_col +1]])
        return df

    def buffered_read_until_empty_row(self, file, skiprows=0, buffer_size=None, **kwargs):
        ''' Reads a file in chunks until an empty row is found. '''
        if buffer_size is None:
            buffer_size = self.filetype_behaviour.get_buffer_size(file)
        start_skipsrows = skiprows
        df = None
        while True:
            if df is None:
                df = self.filetype_behaviour.read_file(file, skiprows=skiprows, nrows=buffer_size, **kwargs)
                headers = df.columns
            else:
                df_next = self.filetype_behaviour.read_file(file, skiprows=skiprows, nrows=buffer_size, header=None, **kwargs)
                if df_next.empty:
                    return df.reset_index().drop(columns=['index'])
                df_next.columns = headers
                for index, row in df_next.iterrows():
                    if row.isnull().all():
                        return pd.concat([df, df_next.iloc[:index]], axis=0).reset_index().drop(columns=['index'])
                df = pd.concat([df, df_next], axis=0)
            # The first loop needs to skip an extra row to account for headers
            if skiprows == start_skipsrows:
                skiprows += buffer_size + 1
            else:
                skiprows += buffer_size
